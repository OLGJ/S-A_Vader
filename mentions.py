import re
import csv
from openpyxl import load_workbook
from stocks import tickers
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer


analyzer = SentimentIntensityAnalyzer()


#Load workbook.
workbook_name = 'work_data_cleaned.xlsx'
wb = load_workbook(workbook_name)


# Sheet titles:
# DD, Discussion, YOLO, News, Earningsthread, Gain, Loss
# Sheet headers:
# Date, ID, Comment, Parent_ID

def count_mentions():
    """
    Loads stocks. Searches comment data for matches, 
    returns percentage of comments who mentions tickers.
    """
    #Mentions
    mention_count = {}
    ticker_count = {}
    
    func_1_score = {}

    #Load tickers and update ticker-frequency-dictionary
    stock_CAPS = tickers()
    for j in stock_CAPS:
        ticker_count[j] = 0
    
    
    #Regex identifier
    stock_mention = r"(\$[a-zA-Z]+)"
    stock_mention = re.compile(stock_mention)
    clean_mention = fr"(?<!\S)(?:{'|'.join(stock_CAPS)})\b"
    
    for each_sheet in wb.worksheets:
        mention_count[each_sheet] = 0
        n_cells = 0
        mentions = 0
       
        for col_cells in each_sheet.iter_cols(min_col=3, max_col=3):
            
            #cell.value = sentence
            for cell in col_cells:
                score = analyzer.polarity_scores(cell.value)['compound']
                if score == 0.0:
                    continue

                #Date of comment.
                target_column = cell.coordinate[1:]
                target_date = 'A'+target_column
                target_date = each_sheet[target_date].value
                #print(target_date)
                #Set to keep track of mentioned stocks in each comment.
                appended = set()
                stock_match = re.findall(stock_mention, cell.value)

                #Check if the $-tick matches
                if stock_match:
                    #Calculate comments compound score with VADER
                    
                    #Potential matches.
                    ticker = [str(x).upper().replace("$", "") for x in stock_match]
                    ticker = list(set(ticker))
                    for tick in ticker:
                        if tick in ticker_count:
                            appended.add(tick)
                            #Add comment´s score and date
                            if tick not in func_1_score:
                                func_1_score[tick] = {target_date : [score]}
                            else:
                                date_check = func_1_score[tick].keys()
                                if target_date not in date_check:
                                    func_1_score[tick][target_date] = [score]
                                else:
                                    func_1_score[tick][target_date].append(score)
                            ticker_count[tick] += 1
        
                #Filters clean matches e.g STOCK, and appends unless already done by the $-match
                clean_match = re.findall(clean_mention, cell.value)
                if clean_match:
                    for j in clean_match:
                        if j not in appended:
                            #Add comment's score and date
                            if j not in func_1_score:
                                func_1_score[j] = {target_date : [score]}
                            else:
                                date_check = func_1_score[j].keys()
                                if target_date not in date_check:
                                    func_1_score[j][target_date] = [score]
                                else:
                                    func_1_score[j][target_date].append(score)
                            ticker_count[j] += 1


    # Relevant stocks:
    # ["GME", "PLTR", "ASO", "TSLA", "AAPL", "APHA", "AMC", "VIAC", "BB", "FUBO", "AMD", "BABA", "NIO"]
    non_relevant = {}
    for tick, score in func_1_score.items():
        x = 0
        for scores in score:
            x += len(score[scores])
        
        non_relevant[tick] = x

    #print({k: v for k, v in sorted(non_relevant.items(), key = lambda item: item[1], reverse=True)})
    
    relevant_stocks = ["GME"]#, "PLTR", "ASO", "TSLA", "AAPL", "APHA", "AMC", "VIAC", "BB", "AMD", "BABA", "NIO", "MVIS", "CLOV"]
    relevant_dictionary = {}
    for tick, score in func_1_score.items():
        if tick in relevant_stocks:
            relevant_dictionary[tick] = score
        
    return relevant_dictionary

def register_score():
    dicton = count_mentions()
    updated_register = []
    #print(dicton)
    counter = {}
    
    #'04/16/21'
    #{Stock : 'GME', 1 : [0.34], 2 : [1, 0.5]}

    for ticker, data in dicton.items():
        
        temp_dict = {}
        #{Stock : ticker, }
        temp_dict['Stock'] = ticker

        keys = data.keys()
        for key in keys:
            new_key = int(key[3:5])-5
            temp_dict[new_key] = data[key]
        #print(temp_dict)
        updated_register.append(temp_dict)           
        x = 0
        for scores in data:
            x += len(data[scores])
        counter[ticker] = x
        
    #updated_register = sorted(updated_register, key=lambda k: k["Date"])
    print({k: v for k, v in sorted(counter.items(), key = lambda item: item[1], reverse=True)})
    #print(updated_register)
    return updated_register

def csv_updater():

    register_scores = register_score()
    #print(register_scores)
    complete_input = []
    prev_vals = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]
    #{Stock, Date, Volume, Price, Score}
    #
    unique = []
    z = 0
    w = 0

    for data_dict in register_scores:
        dictionary_keys = data_dict.keys()
        dictionary_keys = list(dictionary_keys)
        
        stock_key = dictionary_keys.pop(0)
        stock_key = data_dict.get(stock_key)
        #print(dictionary_keys, stock_key)
        for each in dictionary_keys:
            unique.append(each)
        x = 0
        #{1 : [0.35, 1], 2 : [1, 0.5]} - > dictionaryt
        y = 0
        for key in dictionary_keys:
            #data_dict[key] = [0.35, 1]
            x += len(data_dict[key])

            for values in data_dict[key]:
                y += 1
                #print(values)                
                #[0.35] 
                with open(r'C:\Users\Olof\Skola\Kurser\729G40\wallstreet_data_final.csv', 'r') as output_file:

                    csvreader = csv.reader(output_file, delimiter = ',')
                    next(csvreader)

                    for line in csvreader:
                        target_stock = line[0]
                        #print(target_stock)
                        target_date = int(line[1])
                        #print(target_date)
                        if target_stock == stock_key and key == target_date:
                            new_line = line
                            new_line[1] = int(new_line[1])
                            new_line[2] = float(new_line[2])
                            new_line[3] = float(new_line[3])
                            new_line.append(values)
                            complete_input.append(new_line)
                            w += 1
                        #else:
                        #    print(stock_key, values, key)

                    output_file.close()
        z += x

    print('w', w)
    print('z', z)
    complete_input.sort(key = lambda complete_input: complete_input[1])
    
    ctr = 1
    finalized_input = []
    for row in complete_input[1:]:
        row.insert(0, ctr)
        finalized_input.append(row)
        ctr += 1
    
    finalized_input.insert(0, ['Index', 'Stock', 'Date', 'Volume', 'Price', 'Score'])
    #print(finalized_input)
    return finalized_input
    
def csv_writer():
    csv_data = csv_updater()
    #print(csv_data)

    with open(r'C:\Users\Olof\Skola\Kurser\729G40\wsb_modelling_data_gme_test.csv', 'w', newline='') as output_file:
        writer = csv.writer(output_file)
        writer.writerows(csv_data)

        output_file.close()
        


if __name__ == "__main__":
	#count_mentions()
    #register_score()
    #csv_updater()
    csv_writer()