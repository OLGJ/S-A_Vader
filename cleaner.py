#Imports
import re
from openpyxl import load_workbook
from stocks import tickers

#Load workbook.
workbook_name = 'work_data_cp.xlsx'
wb = load_workbook(workbook_name)


# Sheet titles:
# DD, Discussion, YOLO, News, Gain, Loss
# Sheet headers:
# Date, ID, Comment, Parent_ID


def raw_removals():
    """
    Regex cleaner to identify removed, replaced or none content.
    """

    #Regex
    removed = r"(\[borttagen\]|\[raderat\])"

    #Compile
    removed = re.compile(removed)

    #Sheet iteration
    flairs = ["DD", "Discussion", "YOLO", "News", "Gain", "Loss"]
    f_c = 0

    #Iterate
    for each_sheet in wb.worksheets:
        for col_cells in each_sheet.iter_cols(min_col=3, max_col=3):
            for cell in col_cells:
                
                try:
                    cell.value = re.sub(removed, "", str(cell.value))
                except:
                    pass

                if cell.value == None:
                    wb[flairs[f_c]].delete_rows(cell.row)
                    continue
                
                elif len(cell.value) == 0:
                    wb[flairs[f_c]].delete_rows(cell.row)
                    continue
        f_c += 1
    
    #Save changes
    #wb.saveCopyAs(filename=workbook_name)
    wb.save(filename=workbook_name)
    return 

def reddit_unique():
    """
    Regex cleaner for preprocessing different types of expressions unique to Reddit.
    """

    #Regex
    user_reply = r"((\/(u|U)\/+|(u|U)\/+)(?:\w*))"
    subreddit_reply = r"((\/(r|R)\/+|(r|R)\/+)(?:\w*))"

    #Compile
    user_reply = re.compile(user_reply)
    subreddit_reply = re.compile(subreddit_reply)
    
    i = 0
    #iterate
    for each_sheet in wb.worksheets:
        #column 3 = comment column
        for col_cells in each_sheet.iter_cols(min_col=3, max_col=3):
            for cell in col_cells:
                cell.value = re.sub(user_reply, "", str(cell.value))
                cell.value = re.sub(subreddit_reply, "", str(cell.value))

    #Save changes
    wb.save(filename=workbook_name)
    return

def number_and_characters():
    """
    Regex cleaner to remove any digits, and any characters of the following:
    
    $=@&_*#>:`\</{})]|%;~-,([+^”
    
    That are not grouped togheter with another of those characters.
    Intentionally leaves $-characters as they may contain stock mentions.
    """

    #Regex
    reddit_numbers = r"\d+"
    character = r"([\=\@\&\_\*\#\>\:\`\\\<\/\{\}\)\]\|\%\;\~\-\(\[\+\^\”])"
    emoji = r"([\=\@\&\_\*\#\>\:\`\\\<\/\{\}\)\]\|\%\;\~\-\(\[\+\^\”]{2,})"

    #Compile
    reddit_numbers = re.compile(reddit_numbers)
    character = re.compile(character)
    emoji = re.compile(emoji)
      
    #iterate
    for each_sheet in wb.worksheets:
        #column 3 = comment column
        for col_cells in each_sheet.iter_cols(min_col=3, max_col=3):
            for cell in col_cells:
                cell.value = re.sub(reddit_numbers, "", str(cell.value))
                #Remove character if it's not grouped as emoji representation.
                if re.search(character, cell.value) != None: 
                    if re.search(emoji, cell.value) != None:
                        #print('no change:', cell.value)
                        pass
                    else:
                        cell.value = re.sub(character, "", str(cell.value))
                        
    
    #Save changes
    wb.save(filename=workbook_name)
    return

def whitespace_removals():
    """
    Removes any unwanted whitespaces.
    """

    #Regex
    whitespace = r"\s{2,}"
    
    #After running whitespace sub
    endline_whitespace = r"^\s+|\s+$|\s+(?=\s)"

    #Compile
    whitespace = re.compile(whitespace)
    endline_whitespace = re.compile(endline_whitespace)

    #iterate
    for each_sheet in wb.worksheets:
        #column 3 = comment column
        for col_cells in each_sheet.iter_cols(min_col=3, max_col=3):
            for cell in col_cells:

                cell.value = re.sub(whitespace, " ", str(cell.value)) 
                cell.value = re.sub(endline_whitespace, "", str(cell.value))
    
    #Save changes
    wb.save(filename=workbook_name)
    return 

def stats():
    """
        Print statistics once data has been cleaned.
    """
    wb = load_workbook(filename=workbook_name, read_only=True)


    #Sheet iteration
    flairs = ["DD", "Discussion", "YOLO", "News", "Gain", "Loss"]

    n_comments = {"DD": 0, "Discussion": 0, "YOLO": 0, "News": 0, "Gain": 0, "Loss": 0}

    f_c = 0
    for each in wb.worksheets:        
        row_count = each.max_row
        n_comments[flairs[f_c]] += row_count
        f_c += 1
    
    values = n_comments.values()
    total = sum(values)
    print(n_comments)
    print('Totalt:', total)

if __name__ == "__main__":
    raw_removals()
    reddit_unique()
    raw_removals()
    number_and_characters()
    raw_removals()
    whitespace_removals()
    raw_removals()
    stats()